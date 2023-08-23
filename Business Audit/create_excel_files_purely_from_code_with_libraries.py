import os
from openpyxl.styles import Alignment
import xlwings as xw

sys.path.append(os.path.abspath(""))

import SQLFunctions.sql_excel_columns
from CommonLibrary.csv_excel_conversions import *
from CommonLibrary.date_libraries import *


financial_year = '2022 - 2023'

financial_year_folder = financial_year[:4] + ' Jul - ' + financial_year[7:] + ' Jun'

# the number of rows the Excel has. Can edit this in case for some reason, 1000 is not enough
number_of_cells = 1000
quarters_list = [month_period(1), month_period(2), month_period(3), month_period(4)]

# Creating the folders
try:
    year_folder = os.mkdir(financial_year_folder)
except:
    print('Year folder already exists')

# Begin the for loop
auditIDList = [2, 3, 4]

# For counting purposes
iteration_number = 0


for index, quarter in enumerate(quarters_list, start=1):
    # Create the quarter folders
    quarter_folder = os.path.join('./', financial_year_folder, 'Q' + str(index) + ' ' + quarter)
    try:
        os.mkdir(quarter_folder)
    except:
        print('Quarter folder already exists')

    # Creating each Excel Spreadsheet
    for audit_id in auditIDList:
        # /////////////////////////////////////////
        if audit_id == 2:
            spreadsheet_name = "Business Transaction Account"
        elif audit_id == 3:
            spreadsheet_name = "Everyday Offset"
        elif audit_id == 4:
            spreadsheet_name = "Mastercard"
        
        iteration_number = iteration_number + 1
        print('beginning iteration ' + str(iteration_number))

        # Create the directories to dump the CSV files for the Bank Account folders
        csv_data_folder_path = os.path.join(quarter_folder, spreadsheet_name + ' CSV')
        try:
            os.mkdir(csv_data_folder_path)
        except:
            print('CSV Data folder already exists')

        workbook = xl.Workbook()
        workbook.create_sheet(index=0, title='Front Cover')
        workbook.remove(workbook['Sheet'])
        front_cover = workbook.active

        front_cover.row_dimensions[1].height = 46.00

        front_cover['A1'] = spreadsheet_name
        front_cover['A2'] = '  ' + beginning_date(1, financial_year) + ' - ' + ending_date(4, financial_year)

        front_cover.merge_cells('A1:I1')
        front_cover['A1'].font = Font(size=36)

        # making the income and expense sheet creations generic
        data_sheets = ['Income', 'Expenditure']

        for sheet_title_index, sheet_title in enumerate(data_sheets, 1):
            # Create the Income sheet and the inital setups that won't change
            workbook.create_sheet(index=sheet_title_index + 1, title=sheet_title)
            workbook.active = workbook[sheet_title]
            current_sheet = workbook.active
            current_sheet.merge_cells('A1:A3')
            current_sheet.merge_cells('B1:B3')
            current_sheet['A1'] = 'Date'
            current_sheet['B1'] = 'Item Details/Description'
            current_sheet['C1'] = sheet_title + ' Categories'
            current_sheet.column_dimensions['A'].width = 10.55
            current_sheet.column_dimensions['B'].width = 37.09
            current_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_sheet['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_sheet['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            set_borders_all_cells(current_sheet, 1, 3, 1, 1)
            set_borders_all_cells(current_sheet, 1, 3, 2, 2)

            # finding the number of columns in the income/expense section
            if sheet_title_index == 1:
                is_income = True
            elif sheet_title_index == 2:
                is_income = False

            temp_df = SQLFunctions.sql_excel_columns.select_excel_column(audit_id, is_income)
            col_names = list(temp_df["ColumnName"])
            column_number = len(col_names)

            # setting the header rows
            current_sheet.merge_cells(start_row=1, end_row=1, start_column=3, end_column=column_number + 2)
            current_sheet.merge_cells(start_row=1, end_row=3, start_column=column_number + 3,
                                      end_column=column_number + 3)
            current_sheet.merge_cells(start_row=1, end_row=3, start_column=column_number + 4,
                                      end_column=column_number + 4)
            current_sheet.column_dimensions[xl.utils.get_column_letter(column_number + 3)].width = 12.73
            current_sheet.column_dimensions[xl.utils.get_column_letter(column_number + 4)].width = 12.73
            current_sheet.cell(row=1, column=column_number + 3).value = 'Total'
            current_sheet.cell(row=1, column=column_number + 4).value = 'Acc. Total'
            current_sheet.cell(row=1, column=column_number + 3).alignment = Alignment(horizontal='center',
                                                                                      vertical='center', wrap_text=True)
            current_sheet.cell(row=1, column=column_number + 4).alignment = Alignment(horizontal='center',
                                                                                      vertical='center', wrap_text=True)
            set_borders_all_cells(current_sheet, 1, 1, 3, column_number + 2)
            set_borders_all_cells(current_sheet, 1, 3, column_number + 3, column_number + 3)
            set_borders_all_cells(current_sheet, 1, 3, column_number + 4, column_number + 4)

            for column_index, col_name in enumerate(col_names, 1):
                current_sheet.merge_cells(start_row=2, end_row=3, start_column=column_index + 2,
                                          end_column=column_index + 2)
                current_sheet.cell(row=2, column=column_index + 2).value = col_name
                current_sheet.column_dimensions[xl.utils.get_column_letter(column_index + 2)].width = 14
                current_sheet.cell(row=2, column=column_index + 2).alignment = Alignment(horizontal='center',
                                                                                         vertical='center',
                                                                                         wrap_text=True)

            set_borders_all_cells(current_sheet, 1, number_of_cells, 1, 2 + column_number + 2)
            set_number_format(current_sheet, 4, number_of_cells, 3, 2 + column_number + 2)
            set_arial_font(current_sheet, 4, number_of_cells, 3, 2 + column_number + 2)

            frozen_cells = current_sheet['A4']
            current_sheet.freeze_panes = frozen_cells

            total_range = convert_rows_and_columns_to_excel(4, number_of_cells, column_number + 3, column_number + 3)
            count = 3

            # setting the formulae for the ends of the total and acc. total columns
            for row in current_sheet[total_range]:
                count = count + 1
                for cell in row:
                    cell.value = sum_value_formula_excel(False, count, count, 3, column_number + 2)

            acc_total_range = convert_rows_and_columns_to_excel(4, number_of_cells, column_number + 4,
                                                                column_number + 4)
            count = 3
            for row in current_sheet[acc_total_range]:
                count = count + 1
                for cell in row:
                    cell.value = sum_value_formula_excel(False, 4, count, column_number + 3, column_number + 3)

        # making the summary page
        workbook.create_sheet(index=4, title='Summary')
        workbook.active = workbook['Summary']
        summary_sheet = workbook.active
        summary_sheet['A1'] = 'Summary'
        summary_sheet['A2'] = spreadsheet_name
        summary_sheet['A3'] = beginning_date(1, financial_year) + ' - ' + ending_date(4, financial_year)
        summary_sheet['A5'] = 'Bank Balance as at ' + beginning_date(1, financial_year)
        summary_sheet['A7'] = data_sheets[0]

        # income column names and sums
        temp_df_for_income = SQLFunctions.sql_excel_columns.select_excel_column(audit_id, True)
        income_col_names = list(temp_df_for_income["ColumnName"])
        income_column_number = len(income_col_names)

        for income_column_index, col_name in enumerate(income_col_names, 1):
            if 'Transfer' in col_name:
                transfer_index = income_column_index
                continue
            summary_sheet.cell(row=income_column_index + 7, column=1).value = col_name
            summary_sheet.cell(row=income_column_index + 7, column=5).value = sum_value_formula_excel(data_sheets[0], 4,
                                                                                                      number_of_cells,
                                                                                                      income_column_index + 2,
                                                                                                      income_column_index + 2)

        # it's laid out this way since 7 is the number of columns before everything begins being indexed. Can change later if it feels too jank/hard to read
        summary_sheet.cell(row=7 + income_column_number + 1, column=1).value = 'Subtotal'
        summary_sheet.cell(row=7 + income_column_number + 3, column=1).value = income_col_names[transfer_index - 1]
        summary_sheet.cell(row=7 + income_column_number + 5, column=1).value = 'Total'
        summary_sheet.cell(row=7 + income_column_number + 7, column=1).value = data_sheets[1]

        # subtotal sum
        summary_sheet.cell(row=7 + income_column_number + 1, column=5).value = sum_value_formula_excel(False, 8,
                                                                                                       7 + transfer_index - 1,
                                                                                                       5, 5)

        # transfer sum
        summary_sheet.cell(row=7 + income_column_number + 3, column=5).value = sum_value_formula_excel(data_sheets[0],
                                                                                                       4,
                                                                                                       number_of_cells,
                                                                                                       2 + transfer_index,
                                                                                                       2 + transfer_index)

        # income total sum
        summary_sheet.cell(row=7 + income_column_number + 5, column=5).value = '=' + convert_rows_and_columns_to_excel(
            7 + income_column_number + 1, 7 + income_column_number + 1, 5,
            5) + ' + ' + convert_rows_and_columns_to_excel(7 + income_column_number + 3, 7 + income_column_number + 3,
                                                           5, 5)

        # expense column names and sums
        temp_df_for_expense = SQLFunctions.sql_excel_columns.select_excel_column(audit_id, False)
        expense_col_names = list(temp_df_for_expense["ColumnName"])
        expense_column_number = len(expense_col_names)

        for expense_column_index, col_name in enumerate(expense_col_names, 1):
            summary_sheet.cell(row=7 + income_column_number + 7 + expense_column_index, column=1).value = col_name
            summary_sheet.cell(row=7 + income_column_number + 7 + expense_column_index,
                               column=5).value = sum_value_formula_excel(data_sheets[1], 4, number_of_cells,
                                                                         expense_column_index + 2,
                                                                         expense_column_index + 2)

        summary_sheet.cell(row=7 + income_column_number + 7 + expense_column_number + 2, column=1).value = 'Total'
        summary_sheet.cell(row=7 + income_column_number + 7 + expense_column_number + 4,
                           column=1).value = 'Bank Balance as at ' + ending_date(4, financial_year)

        # expense total sum
        summary_sheet.cell(row=7 + income_column_number + 7 + expense_column_number + 2,
                           column=5).value = sum_value_formula_excel(False, 7 + income_column_number + 8,
                                                                     7 + income_column_number + 7 + expense_column_number,
                                                                     5, 5)

        summary_sheet.row_dimensions[1].height = 23.50
        summary_sheet['A1'].font = Font(bold=True, size=18)

        for bold_index in range(5, 7 + income_column_number + 7 + expense_column_number + 5):
            summary_sheet.cell(bold_index, column=1).font = Font(bold=True)

        summary_sheet.column_dimensions['E'].height = 13.73
        summary_sheet.column_dimensions['F'].height = 2.55
        summary_sheet.column_dimensions['H'].height = 11.55

        set_borders_outer_cells(summary_sheet, 8, 7 + income_column_number + 5, 1, 5)
        set_borders_outer_cells(summary_sheet, 7 + income_column_number + 7 + 1,
                                7 + income_column_number + 7 + expense_column_number + 2, 1, 5)
        set_borders_all_cells(summary_sheet, 5, 5, 8, 8)
        set_borders_all_cells(summary_sheet, 7 + income_column_number + 7 + expense_column_number + 4,
                              7 + income_column_number + 7 + expense_column_number + 4, 8, 8)
        set_arial_font(summary_sheet, 5, 5, 8, 8)
        set_arial_font(summary_sheet, 7 + income_column_number + 7 + expense_column_number + 4,
                       7 + income_column_number + 7 + expense_column_number + 4, 8, 8)
        set_arial_font(summary_sheet, 8, 7 + income_column_number + 7 + expense_column_number + 2, 5, 5)
        set_number_format(summary_sheet, 5, 5, 8, 8)
        set_number_format(summary_sheet, 7 + income_column_number + 7 + expense_column_number + 4,
                          7 + income_column_number + 7 + expense_column_number + 4, 8, 8)
        set_number_format(summary_sheet, 8, 7 + income_column_number + 7 + expense_column_number + 2, 5, 5)

        excel_file_name = spreadsheet_name + ' ' + quarter + ' ' + year(index, financial_year) + '.xlsx'
        excel_file_location = os.path.join(quarter_folder, excel_file_name)

        workbook.save(excel_file_location)

        # using xlwings to colour all the Excel sheets white
        workbook = xw.Book(excel_file_location)
        work_sheets = workbook.sheets

        for sheet in work_sheets:
            active_sheet = workbook.sheets[sheet]
            all_data_range = active_sheet.range("1:1048576")
            all_data_range.color = (255, 255, 255)

        workbook.save(excel_file_location)
        workbook.app.quit()
