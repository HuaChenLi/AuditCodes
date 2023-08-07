import openpyxl as xl
from openpyxl.styles import Side, Border, Font


def convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c):
    return xl.utils.get_column_letter(start_c) + str(start_r) + ':' + xl.utils.get_column_letter(end_c) + str(end_r)


def set_borders_all_cells(ws, start_r, end_r, start_c, end_c):
    thin = Side(border_style="thin", color="000000")
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_borders_outer_cells(ws, start_r, end_r, start_c, end_c):
    thin = Side(border_style="thin", color="000000")
    top_row = convert_rows_and_columns_to_excel(start_r, start_r, start_c, end_c)
    bottom_row = convert_rows_and_columns_to_excel(end_r, end_r, start_c, end_c)
    left_column = convert_rows_and_columns_to_excel(start_r, end_r, start_c, start_c)
    right_column = convert_rows_and_columns_to_excel(start_r, end_r, end_c, end_c)
    for cell in ws[top_row][0]:
        cell.border = Border(top=thin)
    for cell in ws[bottom_row][0]:
        cell.border = Border(bottom=thin)
    for row in ws[left_column]:
        for cell in row:
            cell.border = Border(left=thin)
    for row in ws[right_column]:
        for cell in row:
            cell.border = Border(right=thin)
    ws.cell(row=start_r, column=start_c).border = Border(top=thin, left=thin)
    ws.cell(row=start_r, column=end_c).border = Border(top=thin, right=thin)
    ws.cell(row=end_r, column=start_c).border = Border(bottom=thin, left=thin)
    ws.cell(row=end_r, column=end_c).border = Border(bottom=thin, right=thin)


def set_number_format(ws, start_r, end_r, start_c, end_c):
    number_format = '0.00'
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = number_format


def set_arial_font(ws, start_r, end_r, start_c, end_c):
    arial_font = Font(name='Arial', size=10)
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.font = arial_font


def sum_value_formula_excel(sheet_name_title, start_r, end_r, start_c, end_c):
    if isinstance(sheet_name_title, str):
        return '=SUM(' + sheet_name_title + '!' + convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c) + ')'
    elif not sheet_name_title:
        return '=SUM(' + convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c) + ')'
    elif not isinstance(sheet_name_title, str):
        print('first value is not a string')
