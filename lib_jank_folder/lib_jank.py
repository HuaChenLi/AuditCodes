import openpyxl as xl
from openpyxl.styles import Font, Border, Side


def month_period(index):
    if index == 1:
        return    'Jul - Sep'
    elif index == 2:
        return    'Oct - Dec'
    elif index == 3:
        return    'Jan - Mar'
    elif index == 4:
        return    'Apr - Jun'


def year(index, financial_year):
    if index == 1 or index == 2: 
        return     financial_year[:4]
    else:
        return     financial_year[7:]

def beginning_date(index, financial_year):
    if index == 1:
        return '01 Jul ' + year(index, financial_year)
    elif index == 2:
        return '01 Oct ' + year(index, financial_year)
    elif index == 3:
        return '01 Jan ' + year(index, financial_year)
    elif index == 4:
        return '01 Apr ' + year(index, financial_year)

def ending_date(index, financial_year):
    if index == 1:
        return '30 Sep ' + year(index, financial_year)
    elif index == 2:
        return '31 Dec ' + year(index, financial_year)
    elif index == 3:
        return '31 Mar ' + year(index, financial_year)
    elif index == 4:
        return '30 Jun ' + year(index, financial_year)

def convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c):
    return xl.utils.get_column_letter(start_c) + str(start_r) + ':' + xl.utils.get_column_letter(end_c) + str(end_r)


def set_borders_all_cells(ws, start_r, end_r, start_c, end_c):
    thin = Side(border_style="thin", color="000000")
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_borders_outer_cells(ws, start_r, end_r, start_c, end_c):
    thin            = Side(border_style="thin", color="000000")
    top_row         = convert_rows_and_columns_to_excel(start_r, start_r, start_c, end_c)
    bottom_row      = convert_rows_and_columns_to_excel(end_r, end_r, start_c, end_c)    
    left_column     = convert_rows_and_columns_to_excel(start_r, end_r, start_c, start_c)    
    right_column    = convert_rows_and_columns_to_excel(start_r, end_r, end_c, end_c)
    for cell in ws[top_row][0]:
        cell.border = Border(top=thin)
    for cell in ws[bottom_row][0]:
        cell.border = Border(bottom=thin)
    for row in ws[left_column]:
        for cell in row:
            cell.border = Border(left=thin)  
    for row in ws[right_column]:
        for cell in row:
            cell.border = Border(right=thin)
    ws.cell(row = start_r, column = start_c).border         = Border(top=thin, left=thin)
    ws.cell(row = start_r, column = end_c).border           = Border(top=thin, right=thin)
    ws.cell(row = end_r, column = start_c).border           = Border(bottom=thin, left=thin)
    ws.cell(row = end_r, column = end_c).border             = Border(bottom=thin, right=thin)

def set_number_format(ws, start_r, end_r, start_c, end_c):
    number_format               = '0.00'    
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = number_format

def set_arial_font(ws, start_r, end_r, start_c, end_c):
    arial_font                  = Font(name='Arial', size=10)
    cell_range = convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c)
    for row in ws[cell_range]:
        for cell in row:
            cell.font = arial_font

def sum_value_formula_excel(sheet_name_title, start_r, end_r, start_c, end_c):
    if isinstance(sheet_name_title, str):
        return  '=SUM(' + sheet_name_title + '!' +  convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c) + ')'
    elif sheet_name_title == False:
        return  '=SUM(' + convert_rows_and_columns_to_excel(start_r, end_r, start_c, end_c) + ')'
    elif isinstance(sheet_name_title, str) == False:
        print('first value is not a string')
